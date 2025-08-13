from __future__ import annotations

import os
import pickle
from datetime import datetime
from pickletools import optimize
from typing import TYPE_CHECKING

import openpyxl
import yaml
from matplotlib import pyplot as plt
from matplotlib.animation import FuncAnimation, ImageMagickWriter

from .colors import CmdColors

if TYPE_CHECKING:
    from networkx import Graph

    from ..env.env_2d import Env2D
    from .settings import Settings


def create_io_folders() -> None:
    """
    Generate the data and results folders if they do not exist.

    Args:
        None

    Returns:
        None
    """
    if not os.path.exists("data"):
        os.makedirs("data")
        print(
            f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} Data folder "
            f"'data' does not exist. Creating the folder."
        )
    if not os.path.exists("results"):
        os.makedirs("results")
        print(
            f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} Results folder "
            f"'results' does not exist. Creating the folder."
        )


def create_sim_folder() -> tuple[str, str]:
    """
    Generate a unique simulation identifier by combining the current date and a 4-digit
    number. This name is used for both the results folder and as root for the results
    files name.

    Args:
        None

    Returns:
        (sim_id, sim_name) (str, str): simulation id and name.
    """

    # Generate sim name by composing the current date and a 4-digit number
    date = datetime.now().strftime("%Y-%m-%d")
    counter = 0
    while os.path.exists(f"results/{date}-{counter:04}"):
        counter += 1
    sim_id = f"{counter:04}"
    sim_name = f"{date}-{sim_id}"
    os.makedirs(f"results/{sim_name}")  # create sim folder

    # Return the simulation id and name
    return (sim_id, sim_name)


def clean_folder(folder_name: str) -> None:
    """
    Clean the contents of a folder by removing all files and subfolders.

    Args:
        folder_name (str): name of the folder to clean.

    Raise:
        Exception: if there is an error during the deletion of a file or folder.

    Returns:
        None
    """
    # Iterate over the files and subfolders in the folder
    for file in os.listdir(folder_name):
        file_path = os.path.join(folder_name, file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                clean_folder(file_path)
                os.rmdir(file_path)
        except Exception as e:
            print(f"{CmdColors.FAIL}[IO]{CmdColors.ENDC} Error cleaning folder {e}")
            raise e


def load_yaml(file_path: str) -> dict:
    """
    Load data from a YAML file.

    Args:
       file_path (str): path to the YAML file.

    Raise:
        FileNotFoundError: if the file does not exist.
        ValueError: if the file is not formatted correctly.

    Returns:
        data (dict): loaded data dictionary.
    """
    # Verify that folder exists
    if not os.path.exists(file_path):
        print(
            f"{CmdColors.FAIL}[IO]{CmdColors.ENDC} The file {file_path} was "
            "not found."
        )
        raise FileNotFoundError

    # Load the YAML file
    with open(file_path, "r", encoding="utf-8") as file:
        data = yaml.safe_load(file)
        data_type = type(data)
        if data is None:
            print(
                f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} The file {file_path} is "
                "empty. Please check the file."
            )
            data = {}
        elif not data_type == dict:
            print(
                f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} The file {file_path} is "
                "not formatted correctly. Please check the file."
            )
            raise ValueError

    # Return the loaded data
    return data


def log_sim_data(settings: Settings) -> None:
    """
    Log the simulation config and results to a XLSX log file in the results folder.

    Args:
        settings (Settings): Config object with simulation parameters settings.

    Return:
        None
    """
    # Open the log file (create it if it doesn't exist)
    log_path = "results/log.xlsx"
    if os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        sheet = wb["log"]  # select the log sheet
    else:
        wb = openpyxl.Workbook()
        wb.create_sheet("log")  # create the log sheet
        sheet = wb["log"]  # select the log sheet
        print(
            f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} Log file not found. "
            "Creating a new one."
        )

    # Log simulation data
    row = sheet.max_row + 1  # get the first empty row
    sheet.cell(row=row, column=1, value=settings.__start_datetime__.strftime("%Y"))
    sheet.cell(row=row, column=2, value=settings.__start_datetime__.strftime("%m"))
    sheet.cell(row=row, column=3, value=settings.__start_datetime__.strftime("%d"))
    sheet.cell(row=row, column=4, value=settings.sim_id)
    sheet.cell(row=row, column=5, value=settings.env_name)
    sheet.cell(row=row, column=6, value=settings.planner.name)

    # Save the log file
    wb.save(log_path)


def save_sim_data(settings: Settings, graph: Graph) -> None:
    """
    Pickle and compress the simulation settings and the recorded simulation data.

    Args:
        settings (Settings): Config object with simulation parameters settings
        graph (Graph): Graph object generated by the planner

    Returns:
        None
    """
    filename = f"{settings.sim_folder}/{settings.sim_name}-sim-data.pkl"
    data = {"settings": settings, "graph": graph}  # store data in a dictionary
    pickled = pickle.dumps(data)  # dump data dictionary in pickle file
    optimized = optimize(pickled)  # optimize the pickle file
    with open(filename, "wb") as f:
        f.write(optimized)


def write_readme(settings: Settings, env: Env2D) -> None:
    """
    Write README file with most relevant simulation parameters.

    Args:
        settings (Settings): Config object with simulation parameters to save to README.

    Returns:
        None
    """
    filename = f"{settings.sim_folder}/{settings.sim_name}-README.md"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"# README - {settings.sim_name}\n")
        f.write(
            f"Simulation started on {settings.__start_date__} "
            f"at {settings.__start_time__}\n"
        )
        f.write("\n")
        f.write("# World settings:\n")
        f.write(f"-\tWorld origin: [{env.origin[0]}, {env.origin[1]}]\n")
        f.write(f"-\tWorld size: [{env.size[0]}, {env.size[1]}]\n")
        f.write("-\tObstacles vertices:\n")
        for _, obs_vert in enumerate(env.obstacle_vertices):
            f.write(f"\t\t{', '.join(map(str, obs_vert))}\n")
        f.write("-\tGoal vertices:\n")
        for _, goal_vert in enumerate(env.goal_vertices):
            f.write(f"\t\t{', '.join(map(str, goal_vert))}\n")
        f.write("\n")
        f.write("## Initial conditions:\n")
        f.write(
            f"-\tInitial robot location: "
            f"[{env.robot_initial_pos[0]}, {env.robot_initial_pos[1]}]\n"
        )
        f.write("\n")
        f.write("## RRT settings:\n")
        f.write(f"-\tMax edge length: {settings.planner.max_edge_length}\n")
        f.write(f"-\tGoal biasing: {settings.planner.goal_bias}\n")
        f.write(f"-\tGoal bias rate: {settings.planner.goal_bias_rate}\n")
        f.write(
            "-\tTermination when goal reached: "
            f"{settings.planner.goal_reached_termination_condition}\n"
        )
        f.write(
            "-\tTermination at max node number: "
            f"{settings.planner.max_nodes_n_termination_condition}\n"
        )
        f.write(f"-\tMax node number: {settings.planner.max_nodes_n}\n")
        f.write("\n")
        f.write("## Notes:\n")


def save_figure(
    fig: plt.Figure,
    settings: Settings,
    fig_name: str = "fig",
    extension: str = "png",
) -> str:
    """
    Save a figure to the results directory.

    Args:
        fig (plt.Figure): Matplotlib figure object to save.
        settings (Settings): Settings object with simulation parameters and filename.
        fig_name (str): Custom name to the default figure filename (default: 'fig').
        extension (str): Figure extension (default: .png).

    Returns:
        filename (str): The filename of the saved figure.
    """

    # Check extension
    if not extension == settings.plot.format:
        print(
            f"{CmdColors.WARNING}[IO]{CmdColors.ENDC} The input extension "
            f".{extension} does not match the one set in the settings "
            f".{settings.plot.format}. The figure will be saved as a .{extension} "
            "file."
        )

    # Generate filename avoiding duplicates
    counter = 0
    while os.path.exists(
        f"{settings.sim_folder}/"
        f"{settings.sim_name}-{fig_name}-{counter:04}.{extension}"
    ):
        counter += 1
    filename = (
        f"{settings.sim_folder}/"
        f"{settings.sim_name}-{fig_name}-{counter:04}.{extension}"
    )

    # Save figure
    fig.savefig(
        filename,
        dpi=settings.plot.dpi,
        format=extension,
        bbox_inches="tight",
    )

    # Return the figure filename
    return filename


def save_animation(
    anim: FuncAnimation,
    settings: Settings,
    anim_name: str = "anim",
    extension: str = "gif",
) -> str:
    """
    Save an animation to the data directory.

    Args:
        anim (FuncAnimation): FuncAnimation object.
        settings (Settings): Settings object with simulation parameters and filename.
        anim_name (str): custom name for the animation to append at the end of the
                        default filename (default: 'anim').
        extension (str): animation extension (default: .gif).

    Returns:
        filename (str): The filename of the saved animation.
    """

    # Generate filename avoiding duplicates
    counter = 0
    while os.path.exists(
        f"{settings.sim_folder}/"
        f"{settings.sim_name}-{anim_name}-{counter:04}.{extension}"
    ):
        counter += 1
    filename = (
        f"{settings.sim_folder}/"
        f"{settings.sim_name}-{anim_name}-{counter:04}.{extension}"
    )

    # Save animation
    # NOTE: With the pillow writer it does not appear to be possible to loop the gif
    # only 1 time, and by default the gif loops infinitely even with 'repeat=False' set
    # in FuncAnimation. See https://stackoverflow.com/a/71594916/20165241.
    # anim.save(filename, writer="pillow", dpi=300)
    print(f"{CmdColors.OKBLUE}[IO]{CmdColors.ENDC} Saving animation...")
    anim.save(
        filename,
        writer=ImageMagickWriter(fps=24, extra_args=["-loop", "1"]),
        dpi=300,
        progress_callback=print_animation_progress,
    )
    print(f"{CmdColors.OKBLUE}[IO]{CmdColors.ENDC} Animation saved to {filename}.")

    # Return the filename
    return filename


def print_animation_progress(current_frame: int, total_frames: int) -> None:
    """
    Callback to print the status of the animation saving process.

    Args:
        current_frame (int): current frame number.
        total_frames (int): total frames number.

    Returns:
        None
    """
    bar_text = f"{CmdColors.OKBLUE}[IO]{CmdColors.ENDC} Animation save progress:"
    print_progress_bar(
        current_frame,
        total_frames,
        prefix=bar_text,
        suffix="",
    )


def print_progress_bar(
    current_iter: int,
    tot_iter: int,
    prefix: str = "",
    suffix: str = "",
):
    """
    Call in a loop to create terminal progress bar

    Args:
        current_iter (int): current iteration (Int)
        tot_iter (int): total iterations (Int)
        prefix (str): prefix string (Str)
        suffix (str): suffix string (Str)
    """
    # Progress Bar Settings
    fill = "█"
    length = 50
    print_end = "\r"

    # Calculate Progress
    percent = ("{0:." + str(2) + "f}").format(100 * (current_iter / float(tot_iter)))
    filled_length = int(length * current_iter // tot_iter)
    progress_bar = fill * filled_length + "-" * (length - filled_length)

    # Print Progress Bar
    print(f"\r{prefix} |{progress_bar}| {percent}% {suffix}", end=print_end)

    # Print new line on completiojn
    if current_iter == tot_iter:
        print("\n")
